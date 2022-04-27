# !/usr/bin/python
# -*- coding:utf-8 -*-

import pandas as pd
import openpyxl as op

s = r"D:\桌面\111.xlsx"
df = pd.read_excel(s, index_col=0, engine='openpyxl')
print(df)
print(df.columns)
print(df.shape)
wb = op.load_workbook(s)
ws = wb.create_sheet("分期", 1)
ws2 = wb.create_sheet("隶属度", 2)
ws.cell(1, 1, value="年")
ws.cell(1, 2, value="ts")
ws.cell(1, 3, value="te")
for j in range(df.shape[1]):
    for i in range(df.shape[0]):

        if df.iloc[i, j] >= 25:
            ws.cell(j + 2, 1, value=df.columns[j])
            ws.cell(j + 2, 2, value=df.index[i])
            a = i + 1
            break
        ws2.cell(i + 1, j + 1, value=0)
    for i in range(df.shape[0] - 1, -1, -1):
        if df.iloc[i, j] >= 25:
            ws.cell(j + 2, 3, value=df.index[i])
            b = i + 1
            break
        ws2.cell(i + 1, j + 1, value=0)
    for i in range(a, b + 1):
        ws2.cell(i, j + 1, value=1)

ss = s[:-5] + "ceshi.xlsx"
wb.save(ss)
